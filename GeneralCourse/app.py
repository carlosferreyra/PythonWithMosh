from ecommerce.shipping import calc_shipping
from pathlib import Path
import random
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
"""print("Hello World") #Imprime hola mundo
print('*'*10) #Imprime un string de 10 concatenaciones de '*'
x = input("Input :")
print("hello , " + x)
entry = input('>')
words = entry.split(' ')
emojis = {
    ":)": "😊",
    ":(": "☹"

}
output = ''
for word in words:
    output += emojis.get(word, word) + ' '
print(output)


def greet_user(name):
    print(f"Hello {name}")


class Person:
    def __init__(self,aName,aLastName,aAddress):
        self.__Name = aName
        self.__LastName = aLastName
        self.__Address = aAddress
    def GetName(self):
        return self.__Name

    def move(self):
        print("Move")

    def talk(self):
        print("Draw")



for x in range(3):
    random.random()
    random.randint(10,20)
    print(random.choice(["Juan","Pedro","Ana"]))
"""
"""path = Path()
for file in path.glob("*"):
    print(file)
"""
""" Project 1: Automation
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet.cell(1, 1)

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
"""
